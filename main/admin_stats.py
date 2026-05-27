"""
Admin paneli uchun to'liq statistika sahifasi va Excel export.
"""
import io
import json
from datetime import datetime, timedelta

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.db.models import Count, Avg, Q, Max
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Announcement, Course, Survey, TalentedStudentDatabase,
    StateScholarship, BuxduScholarship, BuxduWinnerDatabase,
    Olympiad, BuxduOlympiadWinner, BuxduOlympiad, OakDatabase,
    Conference, DissertationBank, ArticleBank, ResearcherRegulation,
    UserCourseProgress, UserModuleProgress, UserTestResult,
    AssessmentTest, AssessmentTestResult, Literature,
    ScientificSupervisor, SupervisorRequest, OlympiadApplication, OlympiadProgram,
)
from .utils_display import format_assessment_status

User = get_user_model()


# ===== Yordamchi funksiya =====

def _collect_stats():
    """Asosiy statistika ma'lumotlarini yig'adi."""
    now = timezone.now()
    week_ago  = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)
    activity_threshold = now - timedelta(days=20)  # oxirgi 20 kun ichida kirgan = faol

    users_qs = User.objects.all()

    # Foydalanuvchilar
    total_users     = users_qs.count()
    admins_count    = users_qs.filter(Q(is_staff=True) | Q(is_superuser=True)).count()
    regular_users   = total_users - admins_count
    iqtidorli_count = users_qs.filter(assessment_status='iqtidorli').count()
    oddiy_count     = users_qs.filter(assessment_status='oddiy').count()
    # Faollik — oxirgi kirish sanasiga qarab (oxirgi 20 kun)
    active_users    = users_qs.filter(last_login__gte=activity_threshold).count()
    inactive_users  = total_users - active_users
    new_week        = users_qs.filter(date_joined__gte=week_ago).count()
    new_month       = users_qs.filter(date_joined__gte=month_ago).count()

    # Universitetlar bo'yicha
    by_university = list(
        users_qs.exclude(university__isnull=True).exclude(university='')
        .values('university').annotate(c=Count('id')).order_by('-c')[:10]
    )

    # Fakultetlar bo'yicha
    by_faculty = list(
        users_qs.exclude(faculty__isnull=True).exclude(faculty='')
        .values('faculty').annotate(c=Count('id')).order_by('-c')[:15]
    )

    # Fakultet x iqtidorli/oddiy
    faculty_status = list(
        users_qs.exclude(faculty__isnull=True).exclude(faculty='')
        .values('faculty', 'assessment_status')
        .annotate(c=Count('id'))
        .order_by('faculty')
    )

    # Akademik daraja bo'yicha
    by_degree = list(
        users_qs.exclude(academic_degree='').values('academic_degree')
        .annotate(c=Count('id')).order_by('-c')
    )

    # Status (talaba va h.k.)
    by_status = list(
        users_qs.values('status').annotate(c=Count('id')).order_by('-c')
    )

    # Ro'yxatdan o'tish tendensiyasi (oxirgi 30 kun)
    registration_trend = []
    for i in range(29, -1, -1):
        day = (now - timedelta(days=i)).date()
        cnt = users_qs.filter(date_joined__date=day).count()
        registration_trend.append({'date': day.strftime('%d.%m'), 'count': cnt})

    # Testlar
    total_assessment_tests = AssessmentTest.objects.count()
    total_test_results = AssessmentTestResult.objects.count()
    passed_count = AssessmentTestResult.objects.filter(passed=True).count()
    failed_count = total_test_results - passed_count
    avg_score = AssessmentTestResult.objects.aggregate(avg=Avg('percentage'))['avg'] or 0

    # Kontent statistikasi
    content_stats = {
        'E\'lonlar':            Announcement.objects.count(),
        'Kurslar':              Course.objects.count(),
        'So\'rovnomalar':       Survey.objects.count(),
        'Davlat stipendiyalari': StateScholarship.objects.count(),
        'BuxDU stipendiyalari':  BuxduScholarship.objects.count(),
        'Olimpiadalar':         Olympiad.objects.count(),
        'BuxDU olimpiadalari':   BuxduOlympiad.objects.count(),
        'OAK jurnallari':        OakDatabase.objects.count(),
        'Konferensiyalar':       Conference.objects.count(),
        'Dissertatsiyalar':      DissertationBank.objects.count(),
        'Maqolalar':             ArticleBank.objects.count(),
        'Nizomlar':              ResearcherRegulation.objects.count(),
        'Adabiyotlar':           Literature.objects.count(),
        'Iqtidorli baza':        TalentedStudentDatabase.objects.count(),
        'BuxDU g\'oliblari':      BuxduWinnerDatabase.objects.count(),
        'BuxDU olimpiada g\'oliblari': BuxduOlympiadWinner.objects.count(),
    }

    # Kurs jarayoni
    course_progress = {
        'Kurslarda ishtirok':    UserCourseProgress.objects.count(),
        'Modul yakunlangan':     UserModuleProgress.objects.filter(is_completed=True).count() if hasattr(UserModuleProgress, 'is_completed') else UserModuleProgress.objects.count(),
        'Test yechilgan':        UserTestResult.objects.count(),
    }

    # Iqtidor Yo'li — ilmiy rahbarlik
    supervisors_total = ScientificSupervisor.objects.filter(is_active=True).count()
    supervisor_requests_total = SupervisorRequest.objects.count()
    supervisor_pending = SupervisorRequest.objects.filter(status='pending').count()
    supervisor_accepted = SupervisorRequest.objects.filter(status='accepted').count()
    supervisor_rejected = SupervisorRequest.objects.filter(status='rejected').count()
    students_with_supervisor = SupervisorRequest.objects.filter(status='accepted').values('student').distinct().count()

    supervisor_capacity = []
    for sup in ScientificSupervisor.objects.filter(is_active=True).order_by('order', 'full_name'):
        accepted = sup.accepted_count
        supervisor_capacity.append({
            'name': sup.full_name,
            'position': sup.position,
            'accepted': accepted,
            'max': sup.max_students,
            'is_full': accepted >= sup.max_students,
        })

    # Iqtidor Yo'li — olimpiada va volontyor arizalari
    apps_qs = OlympiadApplication.objects.all()
    olympiad_apps_total = apps_qs.count()
    olympiad_apps_new = apps_qs.filter(status='new').count()
    olympiad_apps_reviewed = apps_qs.filter(status='reviewed').count()
    olympiad_apps_approved = apps_qs.filter(status='approved').count()
    olympiad_apps_rejected = apps_qs.filter(status='rejected').count()
    olympiad_type_olympiad = apps_qs.filter(application_type='olympiad').count()
    olympiad_type_volunteer = apps_qs.filter(application_type='volunteer').count()

    apps_by_program = list(
        apps_qs.filter(application_type='olympiad', olympiad__isnull=False)
        .values('olympiad__title')
        .annotate(c=Count('id'))
        .order_by('-c')
    )

    apps_by_status = [
        {'status': 'Yangi', 'c': olympiad_apps_new},
        {'status': "Ko'rib chiqildi", 'c': olympiad_apps_reviewed},
        {'status': 'Tasdiqlandi', 'c': olympiad_apps_approved},
        {'status': 'Rad etildi', 'c': olympiad_apps_rejected},
    ]

    return {
        'now': now,
        'activity_threshold': activity_threshold,
        'total_users': total_users,
        'admins_count': admins_count,
        'regular_users': regular_users,
        'iqtidorli_count': iqtidorli_count,
        'oddiy_count': oddiy_count,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'new_week': new_week,
        'new_month': new_month,
        'by_university': by_university,
        'by_faculty': by_faculty,
        'faculty_status': faculty_status,
        'by_degree': by_degree,
        'by_status': by_status,
        'registration_trend': registration_trend,
        'total_assessment_tests': total_assessment_tests,
        'total_test_results': total_test_results,
        'passed_count': passed_count,
        'failed_count': failed_count,
        'avg_score': round(avg_score, 1),
        'content_stats': content_stats,
        'course_progress': course_progress,
        'supervisors_total': supervisors_total,
        'supervisor_requests_total': supervisor_requests_total,
        'supervisor_pending': supervisor_pending,
        'supervisor_accepted': supervisor_accepted,
        'supervisor_rejected': supervisor_rejected,
        'students_with_supervisor': students_with_supervisor,
        'supervisor_capacity': supervisor_capacity,
        'olympiad_apps_total': olympiad_apps_total,
        'olympiad_apps_new': olympiad_apps_new,
        'olympiad_apps_reviewed': olympiad_apps_reviewed,
        'olympiad_apps_approved': olympiad_apps_approved,
        'olympiad_apps_rejected': olympiad_apps_rejected,
        'olympiad_type_olympiad': olympiad_type_olympiad,
        'olympiad_type_volunteer': olympiad_type_volunteer,
        'apps_by_program': apps_by_program,
        'apps_by_status': apps_by_status,
    }


# ===== Statistika sahifasi =====

@staff_member_required
def statistics_view(request):
    stats = _collect_stats()

    # Chart.js uchun JSON ma'lumotlar
    charts = {
        'user_types': {
            'labels': ['Iqtidorli', 'Oddiy', 'Adminlar'],
            'data': [stats['iqtidorli_count'], stats['oddiy_count'], stats['admins_count']],
        },
        'activity': {
            'labels': ['Faol', 'Faol emas'],
            'data': [stats['active_users'], stats['inactive_users']],
        },
        'faculties': {
            'labels': [f['faculty'] for f in stats['by_faculty']],
            'data':   [f['c'] for f in stats['by_faculty']],
        },
        'universities': {
            'labels': [u['university'] for u in stats['by_university']],
            'data':   [u['c'] for u in stats['by_university']],
        },
        'degrees': {
            'labels': [d['academic_degree'] for d in stats['by_degree']],
            'data':   [d['c'] for d in stats['by_degree']],
        },
        'content': {
            'labels': list(stats['content_stats'].keys()),
            'data':   list(stats['content_stats'].values()),
        },
        'registration_trend': {
            'labels': [d['date'] for d in stats['registration_trend']],
            'data':   [d['count'] for d in stats['registration_trend']],
        },
        'tests': {
            'labels': ['O\'tdi', 'O\'tmadi'],
            'data':   [stats['passed_count'], stats['failed_count']],
        },
        'supervisor_requests': {
            'labels': ['Kutilmoqda', 'Qabul qilindi', 'Rad etildi'],
            'data': [
                stats['supervisor_pending'],
                stats['supervisor_accepted'],
                stats['supervisor_rejected'],
            ],
        },
        'application_status': {
            'labels': [a['status'] for a in stats['apps_by_status']],
            'data': [a['c'] for a in stats['apps_by_status']],
        },
        'application_types': {
            'labels': ['Olimpiada arizalari', 'Volontyor arizalari'],
            'data': [stats['olympiad_type_olympiad'], stats['olympiad_type_volunteer']],
        },
        'applications_by_program': {
            'labels': [p['olympiad__title'] or '—' for p in stats['apps_by_program'][:10]],
            'data': [p['c'] for p in stats['apps_by_program'][:10]],
        },
    }

    return render(request, 'admin/statistics.html', {
        'title': 'Statistika',
        'stats': stats,
        'charts_json': json.dumps(charts, ensure_ascii=False),
    })


def _build_user_iqtidor_maps():
    """Foydalanuvchi bo'yicha ilmiy rahbar va ariza ma'lumotlarini yig'adi."""
    accepted_supervisors = {}
    for req in SupervisorRequest.objects.filter(status='accepted').select_related('supervisor', 'student'):
        accepted_supervisors.setdefault(req.student_id, []).append(req.supervisor)

    latest_supervisor_req = {}
    for req in SupervisorRequest.objects.select_related('supervisor').order_by('student_id', '-created_at'):
        if req.student_id not in latest_supervisor_req:
            latest_supervisor_req[req.student_id] = req

    user_applications = {}
    for app in OlympiadApplication.objects.select_related('olympiad').order_by('user_id', '-created_at'):
        user_applications.setdefault(app.user_id, []).append(app)

    return accepted_supervisors, latest_supervisor_req, user_applications


def _format_user_supervisor_info(user_id, accepted_supervisors, latest_supervisor_req):
    """Excel uchun ilmiy rahbar ma'lumotlarini matn ko'rinishida qaytaradi."""
    supervisors = accepted_supervisors.get(user_id, [])
    if supervisors:
        names = '; '.join(s.full_name for s in supervisors)
        positions = '; '.join(s.position or '—' for s in supervisors)
        emails = '; '.join(s.email or '—' for s in supervisors)
        phones = '; '.join(s.phone or '—' for s in supervisors)
        status = 'Qabul qilindi'
    else:
        names = positions = emails = phones = '—'
        latest = latest_supervisor_req.get(user_id)
        if latest:
            status_map = {'pending': 'Kutilmoqda', 'accepted': 'Qabul qilindi', 'rejected': 'Rad etildi'}
            status = status_map.get(latest.status, latest.status)
            names = latest.supervisor.full_name
            positions = latest.supervisor.position or '—'
            emails = latest.supervisor.email or '—'
            phones = latest.supervisor.phone or '—'
        else:
            status = '—'

    return names, positions, emails, phones, status


def _format_user_applications_info(user_id, user_applications):
    """Excel uchun olimpiada/volontyor arizalarini matn ko'rinishida qaytaradi."""
    apps = user_applications.get(user_id, [])
    if not apps:
        return '—', '—', '—'

    status_labels = dict(OlympiadApplication.STATUS_CHOICES)
    olympiad_parts = []
    volunteer_status = '—'

    for app in apps:
        label = status_labels.get(app.status, app.status)
        if app.application_type == 'volunteer':
            if volunteer_status == '—':
                volunteer_status = label
        else:
            olympiad_parts.append(f"{app.display_title} ({label})")

    olympiad_text = '; '.join(olympiad_parts) if olympiad_parts else '—'
    total_count = len(apps)
    latest = apps[0]
    latest_text = f"{latest.display_title} — {status_labels.get(latest.status, latest.status)}"
    return olympiad_text, volunteer_status, latest_text


# ===== Excel export =====

@staff_member_required
def statistics_excel(request):
    stats = _collect_stats()

    wb = Workbook()

    # Stillar
    BLUE_FILL    = PatternFill('solid', fgColor='3B82F6')
    GREEN_FILL   = PatternFill('solid', fgColor='10B981')
    PURPLE_FILL  = PatternFill('solid', fgColor='8B5CF6')
    ORANGE_FILL  = PatternFill('solid', fgColor='F97316')
    YELLOW_FILL  = PatternFill('solid', fgColor='FACC15')
    RED_FILL     = PatternFill('solid', fgColor='EF4444')
    GREY_FILL    = PatternFill('solid', fgColor='E5E7EB')
    LIGHT_GREEN  = PatternFill('solid', fgColor='D1FAE5')  # Iqtidorli row
    LIGHT_BLUE   = PatternFill('solid', fgColor='DBEAFE')  # Oddiy row
    LIGHT_PURPLE = PatternFill('solid', fgColor='EDE9FE')  # Admin row

    WHITE_FONT = Font(bold=True, color='FFFFFF', size=12)
    BOLD       = Font(bold=True)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def autosize(ws, min_width=10, max_width=50):
        for col_idx, col in enumerate(ws.columns, 1):
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, min_width), max_width)

    # ===== Sheet 1: Umumiy statistika =====
    ws1 = wb.active
    ws1.title = "Umumiy"

    ws1.merge_cells('A1:B1')
    ws1['A1'] = "YOSH TADQIQOTCHI — UMUMIY STATISTIKA"
    ws1['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws1['A1'].fill = PURPLE_FILL
    ws1['A1'].alignment = HEADER_ALIGN
    ws1.row_dimensions[1].height = 28

    ws1['A3'] = "Hisobot sanasi:"
    ws1['B3'] = stats['now'].strftime('%d.%m.%Y %H:%M')
    ws1['A3'].font = BOLD

    rows = [
        ("FOYDALANUVCHILAR",          None, BLUE_FILL),
        ("Jami foydalanuvchilar",     stats['total_users']),
        ("Adminlar (staff/superuser)", stats['admins_count']),
        ("Oddiy foydalanuvchilar",    stats['regular_users']),
        ("Iqtidorli talabalar",        stats['iqtidorli_count']),
        ("Oddiy holatdagilar",         stats['oddiy_count']),
        ("Faol foydalanuvchilar (oxirgi 20 kun)", stats['active_users']),
        ("Faol emas (>20 kun yoki hech kirmagan)", stats['inactive_users']),
        ("Oxirgi 7 kunda yangi",      stats['new_week']),
        ("Oxirgi 30 kunda yangi",     stats['new_month']),
        ("", None),
        ("TESTLAR",                   None, GREEN_FILL),
        ("Jami saralash testlari",     stats['total_assessment_tests']),
        ("Jami test natijalari",      stats['total_test_results']),
        ("O'tgan urunishlar",          stats['passed_count']),
        ("O'tmagan urunishlar",        stats['failed_count']),
        ("O'rtacha ball (%)",          stats['avg_score']),
        ("", None),
        ("KONTENT",                   None, ORANGE_FILL),
    ]

    row_idx = 5
    for r in rows:
        if len(r) == 3 and r[2] is not None:
            # Bo'lim sarlavhasi
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = ws1.cell(row=row_idx, column=1, value=r[0])
            cell.fill = r[2]
            cell.font = WHITE_FONT
            cell.alignment = HEADER_ALIGN
            ws1.row_dimensions[row_idx].height = 22
        elif r[0]:
            ws1.cell(row=row_idx, column=1, value=r[0]).border = border
            ws1.cell(row=row_idx, column=2, value=r[1]).border = border
            ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
        row_idx += 1

    for name, count in stats['content_stats'].items():
        ws1.cell(row=row_idx, column=1, value=name).border = border
        ws1.cell(row=row_idx, column=2, value=count).border = border
        ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
        row_idx += 1

    iqtidor_rows = [
        ("", None),
        ("IQTIDOR YO'LI", None, PURPLE_FILL),
        ("Ilmiy rahbarlar (faol)", stats['supervisors_total']),
        ("Rahbarlik so'rovlari jami", stats['supervisor_requests_total']),
        ("Kutilmoqda", stats['supervisor_pending']),
        ("Qabul qilingan so'rovlar", stats['supervisor_accepted']),
        ("Rad etilgan so'rovlar", stats['supervisor_rejected']),
        ("Rahbar biriktirilgan talabalar", stats['students_with_supervisor']),
        ("", None),
        ("Olimpiada/volontyor arizalari jami", stats['olympiad_apps_total']),
        ("Olimpiada arizalari", stats['olympiad_type_olympiad']),
        ("Volontyor arizalari", stats['olympiad_type_volunteer']),
        ("Yangi arizalar", stats['olympiad_apps_new']),
        ("Ko'rib chiqilgan", stats['olympiad_apps_reviewed']),
        ("Tasdiqlangan", stats['olympiad_apps_approved']),
        ("Rad etilgan", stats['olympiad_apps_rejected']),
    ]
    for r in iqtidor_rows:
        if len(r) == 3 and r[2] is not None:
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = ws1.cell(row=row_idx, column=1, value=r[0])
            cell.fill = r[2]
            cell.font = WHITE_FONT
            cell.alignment = HEADER_ALIGN
            ws1.row_dimensions[row_idx].height = 22
        elif r[0]:
            ws1.cell(row=row_idx, column=1, value=r[0]).border = border
            ws1.cell(row=row_idx, column=2, value=r[1]).border = border
            ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal='right')
        row_idx += 1

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 18

    # ===== Sheet 2: Barcha foydalanuvchilar (rangli) =====
    ws2 = wb.create_sheet("Foydalanuvchilar")
    activity_threshold = stats['activity_threshold']
    accepted_supervisors, latest_supervisor_req, user_applications = _build_user_iqtidor_maps()

    headers = ['#', 'F.I.O', 'Email', 'Telefon', 'Universitet', 'Fakultet',
               'Ilmiy daraja', 'Status', 'Holati', 'Faol (≤20 kun)', 'Admin', 'Ball %',
               'Ilmiy rahbar', 'Rahbar lavozimi', 'Rahbar email', 'Rahbarlik holati',
               'Olimpiada arizalari', 'Volontyor holati', 'Oxirgi kirish', 'Ro\'yxatdan o\'tgan']
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PURPLE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws2.row_dimensions[1].height = 32

    for idx, u in enumerate(User.objects.all().order_by('-date_joined'), start=1):
        row = idx + 1
        is_admin = u.is_staff or u.is_superuser
        is_talented = u.assessment_status == 'iqtidorli'

        # Rang tanlash: admin > iqtidorli > oddiy
        if is_admin:
            fill = LIGHT_PURPLE
        elif is_talented:
            fill = LIGHT_GREEN
        else:
            fill = LIGHT_BLUE

        full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username
        is_active_recent = bool(u.last_login and u.last_login >= activity_threshold)
        sup_name, sup_pos, sup_email, _, sup_status = _format_user_supervisor_info(
            u.id, accepted_supervisors, latest_supervisor_req
        )
        olympiad_text, volunteer_status, _ = _format_user_applications_info(u.id, user_applications)
        row_data = [
            idx,
            full_name,
            u.email or '—',
            u.phone_number or '—',
            u.university or '—',
            u.faculty or '—',
            u.get_academic_degree_display() if u.academic_degree else '—',
            u.status or '—',
            'Iqtidorli' if is_talented else 'Oddiy',
            'Ha' if is_active_recent else 'Yo\'q',
            'Ha' if is_admin else 'Yo\'q',
            round(u.assessment_score or 0, 1),
            sup_name,
            sup_pos,
            sup_email,
            sup_status,
            olympiad_text,
            volunteer_status,
            u.last_login.strftime('%d.%m.%Y %H:%M') if u.last_login else '—',
            u.date_joined.strftime('%d.%m.%Y') if u.date_joined else '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col == 10:
                if is_active_recent:
                    cell.font = Font(bold=True, color='065F46')
                else:
                    cell.font = Font(bold=True, color='991B1B')
            cell.alignment = Alignment(horizontal='center' if col in (1, 9, 10, 11, 12, 16, 18, 19, 20) else 'left',
                                       vertical='center', wrap_text=True)

    autosize(ws2)
    ws2.freeze_panes = 'A2'

    # Legenda
    legend_row = ws2.max_row + 3
    ws2.cell(row=legend_row, column=1, value="RANGLAR IZOHI:").font = BOLD
    ws2.cell(row=legend_row + 1, column=1, value="Admin").fill = LIGHT_PURPLE
    ws2.cell(row=legend_row + 2, column=1, value="Iqtidorli").fill = LIGHT_GREEN
    ws2.cell(row=legend_row + 3, column=1, value="Oddiy").fill = LIGHT_BLUE

    # ===== Sheet 3: Fakultetlar tahlili =====
    ws3 = wb.create_sheet("Fakultetlar")
    headers = ['Fakultet', 'Jami talabalar', 'Iqtidorli', 'Oddiy', 'Iqtidorli %']
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = GREEN_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws3.row_dimensions[1].height = 28

    # Fakultet bo'yicha hisobot
    faculty_map = {}
    for f in stats['faculty_status']:
        name = f['faculty']
        if name not in faculty_map:
            faculty_map[name] = {'iqtidorli': 0, 'oddiy': 0}
        if f['assessment_status'] == 'iqtidorli':
            faculty_map[name]['iqtidorli'] += f['c']
        else:
            faculty_map[name]['oddiy'] += f['c']

    sorted_faculties = sorted(faculty_map.items(), key=lambda x: -(x[1]['iqtidorli'] + x[1]['oddiy']))
    for idx, (fname, counts) in enumerate(sorted_faculties, start=2):
        total = counts['iqtidorli'] + counts['oddiy']
        pct = round((counts['iqtidorli'] / total) * 100, 1) if total else 0
        row_data = [fname, total, counts['iqtidorli'], counts['oddiy'], f"{pct}%"]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=idx, column=col, value=val)
            cell.border = border
            if col == 3:
                cell.fill = LIGHT_GREEN
            elif col == 4:
                cell.fill = LIGHT_BLUE
            cell.alignment = Alignment(horizontal='center' if col != 1 else 'left', vertical='center')

    autosize(ws3)
    ws3.freeze_panes = 'A2'

    # ===== Sheet 4: Universitetlar =====
    ws4 = wb.create_sheet("Universitetlar")
    headers = ['Universitet', 'Foydalanuvchilar soni']
    for col, h in enumerate(headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws4.row_dimensions[1].height = 24
    for idx, u in enumerate(stats['by_university'], start=2):
        ws4.cell(row=idx, column=1, value=u['university']).border = border
        c = ws4.cell(row=idx, column=2, value=u['c'])
        c.border = border
        c.alignment = Alignment(horizontal='center')
    autosize(ws4)

    # ===== Sheet 5: Test natijalari =====
    ws5 = wb.create_sheet("Test natijalari")
    headers = ['#', 'Foydalanuvchi', 'Email', 'Test nomi', 'Ball', 'Foiz', 'O\'tdi', 'Topshirilgan']
    for col, h in enumerate(headers, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.fill = ORANGE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws5.row_dimensions[1].height = 28

    for idx, r in enumerate(AssessmentTestResult.objects.select_related('user', 'assessment_test').order_by('-submitted_at')[:1000], start=2):
        full_name = f"{r.user.first_name or ''} {r.user.last_name or ''}".strip() or r.user.username
        row_data = [
            idx - 1,
            full_name,
            r.user.email or '—',
            r.assessment_test.title if r.assessment_test else '—',
            r.score,
            f"{round(r.percentage, 1)}%",
            'Ha' if r.passed else 'Yo\'q',
            r.submitted_at.strftime('%d.%m.%Y %H:%M') if r.submitted_at else '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws5.cell(row=idx, column=col, value=val)
            cell.border = border
            if col == 7:
                cell.fill = LIGHT_GREEN if r.passed else PatternFill('solid', fgColor='FECACA')
            cell.alignment = Alignment(horizontal='center' if col != 2 and col != 3 and col != 4 else 'left', vertical='center')

    autosize(ws5)
    ws5.freeze_panes = 'A2'

    # ===== Sheet 6: BATAFSIL — har bir foydalanuvchi to'liq ma'lumotlari (fakultet rangida) =====
    ws6 = wb.create_sheet("Batafsil foydalanuvchilar")

    # Fakultet ranglari palitrasi — har bir fakultet uchun yorqin pastel rang
    faculty_palette = [
        'FEF3C7',  # amber
        'D1FAE5',  # emerald
        'DBEAFE',  # blue
        'EDE9FE',  # purple
        'FCE7F3',  # pink
        'FEE2E2',  # red
        'CFFAFE',  # cyan
        'ECFCCB',  # lime
        'FED7AA',  # orange
        'E0E7FF',  # indigo
        'F3E8FF',  # violet
        'F0FDF4',  # light green
        'FFE4E6',  # rose
        'E0F2FE',  # sky
        'FFEDD5',  # peach
        'FAE8FF',  # fuchsia
        'F0FDFA',  # teal
        'FFF7ED',  # warm
        'F5F3FF',  # lavender
        'FFFBEB',  # cream
    ]
    faculty_color_map = {}
    no_faculty_color = 'F3F4F6'  # kulrang — fakulteti yo'qlar uchun

    def get_faculty_color(fname):
        if not fname or not fname.strip():
            return no_faculty_color
        if fname not in faculty_color_map:
            color = faculty_palette[len(faculty_color_map) % len(faculty_palette)]
            faculty_color_map[fname] = color
        return faculty_color_map[fname]

    detail_headers = [
        '#', 'Fakultet', 'F.I.O', 'Username', 'Email', 'Telefon',
        'Yashash xududi', 'Universitet', 'Ilmiy daraja', 'Status',
        'Holati (Iqtidorli/Oddiy)', 'Saralash balli (%)',
        'Test urinishlari', 'O\'tgan urinishlar', 'Eng yuqori ball (%)',
        'Saralash topshirilgan', 'Keyingi urinish',
        'Faol (≤20 kun)', 'Staff', 'Superuser',
        'Ro\'yxatdan o\'tgan', 'Oxirgi kirish',
        'Ilmiy rahbar', 'Rahbar lavozimi', 'Rahbar email', 'Rahbar telefon', 'Rahbarlik holati',
        'Olimpiada arizalari', 'Volontyor holati', 'Oxirgi ariza',
    ]
    for col, h in enumerate(detail_headers, 1):
        c = ws6.cell(row=1, column=col, value=h)
        c.fill = PURPLE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws6.row_dimensions[1].height = 38

    # Foydalanuvchilarni avval fakultet bo'yicha guruhlab, fakultet ichida F.I.O bo'yicha tartiblaymiz
    all_users = User.objects.all().order_by('faculty', 'last_name', 'first_name')
    accepted_supervisors, latest_supervisor_req, user_applications = _build_user_iqtidor_maps()

    for idx, u in enumerate(all_users, start=1):
        row = idx + 1
        faculty_name = (u.faculty or '').strip()
        color = get_faculty_color(faculty_name)
        fill = PatternFill('solid', fgColor=color)

        full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username

        # Test statistikasi
        test_results = AssessmentTestResult.objects.filter(user=u)
        attempts_count = test_results.count()
        passed_count_u = test_results.filter(passed=True).count()
        best_score = test_results.aggregate(Max('percentage'))['percentage__max'] or 0

        is_talented = u.assessment_status == 'iqtidorli'
        is_active_recent = bool(u.last_login and u.last_login >= stats['activity_threshold'])

        sup_name, sup_pos, sup_email, sup_phone, sup_status = _format_user_supervisor_info(
            u.id, accepted_supervisors, latest_supervisor_req
        )
        olympiad_text, volunteer_status, latest_app_text = _format_user_applications_info(
            u.id, user_applications
        )

        row_data = [
            idx,
            faculty_name or '—',
            full_name,
            u.username or '—',
            u.email or '—',
            u.phone_number or '—',
            u.residence_region or '—',
            u.university or '—',
            u.get_academic_degree_display() if u.academic_degree else '—',
            u.status or '—',
            'Iqtidorli' if is_talented else 'Oddiy',
            round(u.assessment_score or 0, 1),
            attempts_count,
            passed_count_u,
            round(best_score, 1),
            u.assessment_taken_at.strftime('%d.%m.%Y %H:%M') if u.assessment_taken_at else '—',
            u.assessment_next_attempt.strftime('%d.%m.%Y %H:%M') if u.assessment_next_attempt else '—',
            'Ha' if is_active_recent else 'Yo\'q',
            'Ha' if u.is_staff else 'Yo\'q',
            'Ha' if u.is_superuser else 'Yo\'q',
            u.date_joined.strftime('%d.%m.%Y %H:%M') if u.date_joined else '—',
            u.last_login.strftime('%d.%m.%Y %H:%M') if u.last_login else '—',
            sup_name,
            sup_pos,
            sup_email,
            sup_phone,
            sup_status,
            olympiad_text,
            volunteer_status,
            latest_app_text,
        ]
        center_cols = {1, 12, 13, 14, 15, 18, 19, 20, 25}
        for col, val in enumerate(row_data, 1):
            cell = ws6.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col == 11:
                cell.font = Font(bold=True, color='065F46' if is_talented else '1E3A8A')
            if col == 18:
                cell.font = Font(bold=True, color='065F46' if is_active_recent else '991B1B')
            if col in center_cols:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    autosize(ws6, min_width=12, max_width=42)
    ws6.freeze_panes = 'C2'

    # ----- Fakultetlar ranglar izohi -----
    if faculty_color_map:
        legend_row = ws6.max_row + 3
        title_cell = ws6.cell(row=legend_row, column=1, value="FAKULTETLAR RANGLAR IZOHI:")
        title_cell.font = Font(bold=True, size=12, color='8B5CF6')
        legend_row += 1

        # 2 ustunda joylashtiramiz
        legend_items = list(faculty_color_map.items())
        for i, (fname, color) in enumerate(legend_items):
            r = legend_row + (i // 2)
            c_start = 1 if i % 2 == 0 else 4
            color_cell = ws6.cell(row=r, column=c_start, value="")
            color_cell.fill = PatternFill('solid', fgColor=color)
            color_cell.border = border
            name_cell = ws6.cell(row=r, column=c_start + 1, value=fname)
            name_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal='left', vertical='center')
            # Sanab ko'rsatish
            count_cell = ws6.cell(
                row=r, column=c_start + 2,
                value=User.objects.filter(faculty=fname).count()
            )
            count_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fakulteti yo'qlar uchun ham
        no_fac_count = User.objects.filter(Q(faculty__isnull=True) | Q(faculty='')).count()
        if no_fac_count:
            r = legend_row + (len(legend_items) // 2) + 1
            ws6.cell(row=r, column=1).fill = PatternFill('solid', fgColor=no_faculty_color)
            ws6.cell(row=r, column=1).border = border
            ws6.cell(row=r, column=2, value="(Fakulteti kiritilmagan)").font = Font(italic=True)
            ws6.cell(row=r, column=3, value=no_fac_count).alignment = Alignment(horizontal='center')

    # ===== Sheet 7: Ilmiy rahbarlik so'rovlari =====
    ws7 = wb.create_sheet("Rahbarlik so'rovlari")
    req_headers = ['#', 'Talaba', 'Email', 'Universitet', 'Fakultet', 'Rahbar', 'Lavozim',
                   'Holat', 'Yuborilgan', 'Javob vaqti', 'Izoh']
    for col, h in enumerate(req_headers, 1):
        c = ws7.cell(row=1, column=col, value=h)
        c.fill = PURPLE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws7.row_dimensions[1].height = 28

    status_fill_map = {
        'pending': PatternFill('solid', fgColor='FEF3C7'),
        'accepted': PatternFill('solid', fgColor='D1FAE5'),
        'rejected': PatternFill('solid', fgColor='FEE2E2'),
    }
    status_label_map = {'pending': 'Kutilmoqda', 'accepted': 'Qabul qilindi', 'rejected': 'Rad etildi'}

    for idx, req in enumerate(
        SupervisorRequest.objects.select_related('student', 'supervisor').order_by('-created_at'), start=2
    ):
        u = req.student
        full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username
        row_data = [
            idx - 1,
            full_name,
            u.email or '—',
            u.university or '—',
            u.faculty or '—',
            req.supervisor.full_name,
            req.supervisor.position or '—',
            status_label_map.get(req.status, req.status),
            req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else '—',
            req.decided_at.strftime('%d.%m.%Y %H:%M') if req.decided_at else '—',
            req.decision_reason or '—',
        ]
        row_fill = status_fill_map.get(req.status, GREY_FILL)
        for col, val in enumerate(row_data, 1):
            cell = ws7.cell(row=idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = LEFT_ALIGN if col not in (1, 8) else HEADER_ALIGN
    autosize(ws7)
    ws7.freeze_panes = 'A2'

    # ===== Sheet 8: Olimpiada arizalari =====
    ws8 = wb.create_sheet("Olimpiada arizalari")
    app_headers = ['#', 'F.I.O', 'Email', 'Telefon', 'Universitet', 'Fakultet',
                   'Holati (iqtidor)', 'Olimpiada', 'Ariza turi', 'Motivatsiya',
                   'Holat', 'Yuborilgan', 'Admin izohi']
    for col, h in enumerate(app_headers, 1):
        c = ws8.cell(row=1, column=col, value=h)
        c.fill = ORANGE_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws8.row_dimensions[1].height = 28

    app_status_fill = {
        'new': PatternFill('solid', fgColor='DBEAFE'),
        'reviewed': PatternFill('solid', fgColor='FEF3C7'),
        'approved': PatternFill('solid', fgColor='D1FAE5'),
        'rejected': PatternFill('solid', fgColor='FEE2E2'),
    }

    for idx, app in enumerate(
        OlympiadApplication.objects.select_related('user', 'olympiad').order_by('-created_at'), start=2
    ):
        u = app.user
        full_name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username
        row_data = [
            idx - 1,
            full_name,
            u.email or '—',
            u.phone_number or '—',
            u.university or '—',
            u.faculty or '—',
            format_assessment_status(u),
            app.display_title,
            app.get_application_type_display(),
            app.motivation or '—',
            app.get_status_display(),
            app.created_at.strftime('%d.%m.%Y %H:%M') if app.created_at else '—',
            app.admin_note or '—',
        ]
        row_fill = app_status_fill.get(app.status, GREY_FILL)
        for col, val in enumerate(row_data, 1):
            cell = ws8.cell(row=idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = LEFT_ALIGN
    autosize(ws8)
    ws8.freeze_panes = 'A2'

    # ===== Sheet 9: Ilmiy rahbarlar sig'imi =====
    ws9 = wb.create_sheet("Ilmiy rahbarlar")
    sup_headers = ['#', 'F.I.O', 'Lavozim', 'Mutaxassislik', 'Email', 'Telefon',
                   'Qabul qilingan', 'Maksimum', 'Bo\'sh joy', 'To\'liq', 'Faol']
    for col, h in enumerate(sup_headers, 1):
        c = ws9.cell(row=1, column=col, value=h)
        c.fill = GREEN_FILL
        c.font = WHITE_FONT
        c.alignment = HEADER_ALIGN
        c.border = border
    ws9.row_dimensions[1].height = 28

    for idx, sup in enumerate(ScientificSupervisor.objects.all().order_by('order', 'full_name'), start=2):
        accepted = sup.accepted_count
        free_slots = max(sup.max_students - accepted, 0)
        is_full = accepted >= sup.max_students
        row_data = [
            idx - 1,
            sup.full_name,
            sup.position or '—',
            sup.specialty or '—',
            sup.email or '—',
            sup.phone or '—',
            accepted,
            sup.max_students,
            free_slots,
            'Ha' if is_full else 'Yo\'q',
            'Ha' if sup.is_active else 'Yo\'q',
        ]
        row_fill = PatternFill('solid', fgColor='FEE2E2') if is_full else PatternFill('solid', fgColor='D1FAE5')
        for col, val in enumerate(row_data, 1):
            cell = ws9.cell(row=idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = HEADER_ALIGN if col in (1, 7, 8, 9, 10, 11) else LEFT_ALIGN
    autosize(ws9)
    ws9.freeze_panes = 'A2'

    # ===== Saqlash =====
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"statistika_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
