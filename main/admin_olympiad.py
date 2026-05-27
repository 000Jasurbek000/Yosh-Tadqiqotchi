"""Olimpiada arizalarini Excel formatda export qilish."""
import io
from datetime import datetime

from django.http import HttpResponse

from .models import OlympiadApplication
from .utils_display import format_assessment_status


STATUS_COLORS = {
    'new':      'DBEAFE',   # ko'k
    'reviewed': 'FEF3C7',   # sariq
    'approved': 'D1FAE5',   # yashil
    'rejected': 'FEE2E2',   # qizil
}


def generate_applications_excel(queryset=None):
    """OlympiadApplication queryset uchun Excel fayl yaratadi."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan: pip install openpyxl", status=500)

    if queryset is None:
        queryset = OlympiadApplication.objects.all()
    queryset = queryset.select_related('user', 'olympiad').order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "Olimpiada arizalari"

    # ─── Sarlavha qatori ───
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='9CA3AF')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = [
        '#', 'F.I.O', 'Email', 'Telefon', 'Universitet', 'Fakultet',
        'Daraja', 'Status (iqtidor)', 'Olimpiada', 'Motivatsiya',
        'Holat', 'Yuborilgan vaqt', 'Admin izohi'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # ─── Ma'lumot qatorlari ───
    data_align = Alignment(vertical='center', wrap_text=True)
    for row_num, app in enumerate(queryset, 2):
        u = app.user
        ws.cell(row=row_num, column=1,  value=app.id)
        ws.cell(row=row_num, column=2,  value=u.get_full_name() or u.username or '—')
        ws.cell(row=row_num, column=3,  value=u.email or '—')
        ws.cell(row=row_num, column=4,  value=u.phone_number or '—')
        ws.cell(row=row_num, column=5,  value=u.university or '—')
        ws.cell(row=row_num, column=6,  value=getattr(u, 'faculty', '') or '—')
        ws.cell(row=row_num, column=7,  value=u.get_academic_degree_display() if u.academic_degree else '—')
        ws.cell(row=row_num, column=8,  value=format_assessment_status(u))
        ws.cell(row=row_num, column=9,  value=app.display_title)
        ws.cell(row=row_num, column=10, value=app.motivation or '—')
        ws.cell(row=row_num, column=11, value=app.get_status_display())
        ws.cell(row=row_num, column=12, value=app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else '—')
        ws.cell(row=row_num, column=13, value=app.admin_note or '—')

        # status rangini butun qatorga qo'llash
        fill_color = STATUS_COLORS.get(app.status, 'FFFFFF')
        row_fill = PatternFill('solid', fgColor=fill_color)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_num, column=col)
            c.fill = row_fill
            c.alignment = data_align
            c.border = border

    # ─── Ustun kengligi ───
    widths = [5, 28, 28, 18, 28, 24, 14, 18, 36, 40, 16, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    # ─── Legend (status ranglari) ───
    legend_row = ws.max_row + 3
    ws.cell(row=legend_row, column=1, value='Holat ranglari:').font = Font(bold=True, size=11)
    legend = [
        ('Yangi',         'DBEAFE'),
        ("Ko'rib chiqildi", 'FEF3C7'),
        ('Tasdiqlandi',   'D1FAE5'),
        ('Rad etildi',    'FEE2E2'),
    ]
    for i, (label, color) in enumerate(legend, 1):
        c = ws.cell(row=legend_row + i, column=1, value=label)
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = border

    # ─── Statistika varag'i ───
    ws2 = wb.create_sheet("Statistika")
    ws2.cell(row=1, column=1, value='Olimpiada bo\'yicha arizalar statistikasi').font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value='Olimpiada').font = Font(bold=True)
    ws2.cell(row=3, column=2, value='Jami').font = Font(bold=True)
    ws2.cell(row=3, column=3, value='Yangi').font = Font(bold=True)
    ws2.cell(row=3, column=4, value="Ko'rib chiqildi").font = Font(bold=True)
    ws2.cell(row=3, column=5, value='Tasdiqlandi').font = Font(bold=True)
    ws2.cell(row=3, column=6, value='Rad etildi').font = Font(bold=True)
    for c in range(1, 7):
        ws2.cell(row=3, column=c).fill = header_fill
        ws2.cell(row=3, column=c).font = Font(bold=True, color='FFFFFF')

    from .models import OlympiadProgram
    row = 4
    for prog in OlympiadProgram.objects.all():
        apps_qs = OlympiadApplication.objects.filter(olympiad=prog, application_type='olympiad')
        ws2.cell(row=row, column=1, value=prog.title)
        ws2.cell(row=row, column=2, value=apps_qs.count())
        ws2.cell(row=row, column=3, value=apps_qs.filter(status='new').count())
        ws2.cell(row=row, column=4, value=apps_qs.filter(status='reviewed').count())
        ws2.cell(row=row, column=5, value=apps_qs.filter(status='approved').count())
        ws2.cell(row=row, column=6, value=apps_qs.filter(status='rejected').count())
        row += 1

    # Volontyor arizalari
    vol_qs = OlympiadApplication.objects.filter(application_type='volunteer')
    ws2.cell(row=row, column=1, value=OlympiadApplication.VOLUNTEER_TITLE)
    ws2.cell(row=row, column=2, value=vol_qs.count())
    ws2.cell(row=row, column=3, value=vol_qs.filter(status='new').count())
    ws2.cell(row=row, column=4, value=vol_qs.filter(status='reviewed').count())
    ws2.cell(row=row, column=5, value=vol_qs.filter(status='approved').count())
    ws2.cell(row=row, column=6, value=vol_qs.filter(status='rejected').count())

    for i, w in enumerate([42, 10, 10, 18, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ─── Faylga yozish va qaytarish ───
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"olimpiada_arizalari_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def olympiad_applications_excel(request):
    """Admin URL — barcha olimpiada arizalarini Excel formatda yuklab olish."""
    return generate_applications_excel()
